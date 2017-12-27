# -*- coding:utf-8 -*-
#-------------------------------------------------------------------------------
# Name:        自动整理补丁
# Purpose:     支持 2007及以后的 xecel 版本
#
# Author:      zou.yw
#
# Created:     25/10/2017
# Modifiy:     14/11/2017
# Copyright:   (c) zou.yw 2017
# Licence:     GPL
#-------------------------------------------------------------------------------


from openpyxl import Workbook, load_workbook
import operator
import os
import re


def patch_check(path='./', excel_file=None):
    """
    功能：待验证补丁检查，通过待验证补丁文档内的内容与本地实际文件做对比
    参数：
        入参：path=补丁目录（如果没有则默认当前目录）;
                  execel_file=表格文件路径（可选）如果没有表格路径，则读取目录中的第一个 .xlsx 文件
    """
    if not excel_file:
        for i in os.listdir(path):
            if os.path.splitext(i)[1].lower() == '.xlsx':
                if path[-1] != '/' or path[-1] != '\\':
                    path += '/'
                excel_file = path + i
                break
    dll = sql = list()
    dll,sql = read_file_name_from_update_checkin_execl(excel_file)
    programe_file = ('.dll', '.exe', '.lib')
    script_file = ('.sql', '.rps')
    # 不检查的文件，如：版本号更新文件，视图刷新脚本    
    no_check = ('版本脚本.sql', '视图刷新脚本.sql')
    # dll文件存放在同一个文件夹，无需考虑重名问题
    # 去重
    dll_tmp = set(dll)
    dll = list(dll_tmp)
    dll.sort()
    dll_for_check = dll.copy()
    for t in range(len(dll_for_check)):
        # 有部分人员签入不喜欢带上后缀，exe、lib 等
        dll_for_check[t] = data_format(dll_for_check[t])
        if dll_for_check[t][-4:] in programe_file:
            dll_for_check[t] = dll_for_check[t][0:-4]
    
    sql.sort()
    for i in range(sql.count('')):
        sql.remove('')
    sql_for_check = sql.copy()
    for t in range(len(sql_for_check)):
        sql_for_check[t] = data_format(sql_for_check[t])
        if sql_for_check[t][-4:] not in script_file:
            sql_for_check[t] += '.sql' 
    
    # 本地文件遍历
    local_file = list()
    local_dll_list = list()
    local_sql_list = list()
    for k in ('程序', '报表', '脚本', '人力web'):
        local_file.extend(get_file_list(path + k))        
    for t in local_file:
        first_name, last_name = os.path.splitext(t)
        if last_name in programe_file:
            # 程序文件取不带后缀的文件名进行比较
            local_dll_list.append(data_format(first_name))
        elif last_name in script_file:
            local_sql_list.append(data_format(t))
    local_dll_list.sort()
    local_sql_list.sort()
    for k in no_check:
        # 有可能有多个位置存放了不需要检查的脚本（如：脚本更新.sql）
        for t in range(local_sql_list.count(k)):
            local_sql_list.remove(k)
    
    dll_miss_in_excel = list()
    if operator.eq(dll_for_check, local_dll_list):
        print('dll is eq')
        # 返回的列表是文档内存在，但是本地文件不存在的
        dll = None
    else: 
        for t in local_dll_list:
            # excel 中的元素去掉本地找到的列表元素
            # print("lcd:", local_dll_list)
            # print("ecd:", dll_for_check)
            try:
                index_t = dll_for_check.index(t)
                dll.pop(index_t)
                dll_for_check.pop(index_t)
            except ValueError:
                dll_miss_in_excel.append(t)
            except Exception as bug:
                print("we have a bug", bug)
        print("excel文件中存在，在本地没有找到的dll：", len(dll), dll)
        print("本地找到，excel中没有的dll：", len(dll_miss_in_excel), dll_miss_in_excel)
        
    sql_miss_in_excel = list()
    if operator.eq(sql_for_check, local_sql_list):
        print('sql is eq')
        sql = None
    else:
        sql_tmp = list()        
        for k in local_sql_list:
            # excel 中的元素去掉本地找到的列表元素      
            # print("lcs:", len(local_sql_list), local_sql_list)
            # print("ecs:", len(sql_for_check), sql_for_check)
            try:
                index_t = sql_for_check.index(k)
                sql_tmp.append(sql.pop(index_t))
                sql_for_check.pop(index_t)
            except ValueError:
                sql_miss_in_excel.append(k)
            except Exception as bug:
                print("wo have a bug:", bug)                
        print("excel文件中存在，在本地没有找到的脚本：", sql)
        print("本地找到，excel中没有的脚本：", sql_miss_in_excel)
    # 返回的内容包括：本地缺失的dll、excel缺失的dll、本地确实的sql、excel缺失的sql
    return (dll, dll_miss_in_excel, sql, sql_miss_in_excel)


def read_file_name_from_update_checkin_execl(excel_file=None):
    """
    功能：待验证补丁表格文档读取
    入参：表格文件路径
    返回：提取出文件中包含的所有文件（报表、dll、sql 等）的2个集合（2个返回结果）
    """
    dll_set = list()
    script_set = list()
    sql_set_tmp = set()
    excel_data = read_excel(data=excel_file, read_type='row')
    if not excel_data:
        print("read excel err")
        return 0
    # print(excel_data)
    dll_index = 0
    sql_index = 3
    # 表头特征
    # tabel_head = (r'修改日期', r'对应DLL文件', r'SQL脚本/报表/其它配置文件(含路径')
    # 遍历表格，替换字符串，方便去重
    # 遍历每个工作簿
    for key in excel_data:
        # 遍历字典中的每个值（嵌套的列表），等同于excel的一张工作表
        sheet_data = excel_data[key]
        # print(sheet_data)
        # 遍历每个列表（等同于excel的每一行/每一列）
        for cells in sheet_data:
            if '修改日期' in cells:
                try:
                    dll_index = cells.index(r'对应DLL文件')
                    sql_index = cells.index(r'SQL脚本/报表/其它配置文件(含路径)')
                except:
                    pass
            else:
                dll_tmp = data_format(cells[dll_index]).split('\n')
                if dll_tmp != ['']:
                    dll_set.extend(dll_tmp)
                script_tmp = data_format(cells[sql_index]).split('\n')
                if script_tmp != ['']:
                    script_set.extend(script_tmp)                    
    # 去重，把相同路径下同名文件合并
    script_tmp = list(set(script_set))
    script_set_tmp = set()    
    for i in script_tmp:
        t = i.split('hkdatabase')[-1]
        script_set_tmp.add(t)
    # 要把脚本、报表的服务器路径剔除
    script_set = []
    for i in script_set_tmp:
        t = i.split('/')[-1]
        if t != ['']: 
            script_set.append(t) 
    # dll_set.remove('')
    # script_set.remove('')
    return dll_set, script_set


def read_excel(data=None, read_type='row',pack="dict"):
    """
    参数：
        data=excel文件路径，read_type=row/column按行读还是按列读，
        pack=dict/set/list 返回结果是字典还是集合（如果是集合，所有的工作簿单元格都放到一个集合中）
    功能：遍历excel文件，返回数据列表或集合（默认按行）
    返回：包含元组的字典（每个工作簿作为一个字典key，元组内容为一个2纬列表）
              egg: {"key1":[[v1,v1-1], [v2,v2-1], [...]], "key2":[[...]], ...}
    备注：按列读取有BUG，取出来的数据不全，而且没有了 key
    """

    try:
        # 打开文件
        workbook = load_workbook(data)
        # 获取所有sheet，返回列表，格式：[u'sheet1', u'sheet2']
        workbook_sheets_list = workbook.get_sheet_names()
        wb_sheets = {}
        data_set = set()
        data_list = list()
        for sheet_name in workbook_sheets_list:
            sheet = workbook.get_sheet_by_name(sheet_name)
            # 统计空行的数量，如果连续空一定数量行，则认为表格后面都是空的，跳过
            # 避免有些单元格设置了格式，但是没有实际数据时，会一直在遍历这些单元格
            null_cnt = 0
            sheet_all_rows = []
            if read_type =='row':
                sheet_item = sheet.rows
            else:
                sheet_item = sheet.columns
            # 遍历工作簿所有单元格
            for colu_row in sheet_item:
                # 遍历每行或列
                tmp = []
                # 遍历每个单元格
                for cel_index in range(len(colu_row)):
                    if colu_row[cel_index].value:
                        null_cnt = 0                                           
                        tmp.append(colu_row[cel_index].value)
                        data_set.add(colu_row[cel_index].value)
                        data_list.append(colu_row[cel_index].value)
                    else:
                        tmp.append("")
                        data_set.add("")
                        data_list.append("")
                    # print(sheet_all_rows)
                    # for cell in colu_row:
                    null_cnt +=1
                sheet_all_rows.append(tmp)

                # 连续10行空
                if null_cnt > 10:
                    break
            wb_sheets[sheet.title] = sheet_all_rows
    except Exception as err:
        print(err)
        return None
    if pack == "dict":
        return wb_sheets
    elif pack == "set":
        return data_set
    else:
        return data_list


def write_excel(data=None, file_name='./new_excel.xlsx'):
    """
    功能：把数据写入 excel
    入参：data=要写入的数据（支持 列表、元组、字符串）; file_name=要写入的文件
    返回值：写入成功 Ture; 写入失败 False
    """

    if not isinstance(data,list):
        # return "The data to write , not a list"
        return None
    # 新建一个工作簿（内存中），用来存放输出结果
    wb_new = Workbook()
    # 新建一张表
    ws_new = wb_new.active
    try:
        ws_new.append(data)
    except  Exception as bug:
        ws_new.append(['we have a problme. i have a bug'])
        ws_new.append([bug])
    finally:
        try:
            wb_new.save(file_name)
        except  Exception as err:
            print('无法保存文件，或许在编辑中')
    return


def data_format(data):
    """
    功能：对数据做格式化处理
    """
    try:
        if isinstance(data, list):
            for i in range(len(data)):
                tmp_data = str(data)
                tmp_data = tmp_data.strip()
                tmp_data = data[i].lower()
                tmp_data = tmp_data.replace("\\","/")                
                data[i] = tmp_data
        else:
            tmp_data = str(data)
            tmp_data = tmp_data.lower()
            tmp_data = tmp_data.strip()
            tmp_data = tmp_data.replace("\\","/")            
            data = tmp_data
        return data
    except Exception as err:
        print(err)
        return None

def get_file_list(file_path=None):
    """获取文件夹列表"""
    if not file_path:
        return None
    # file_set = set()
    file_list = []
    try:
        for fpath, dirs, fs in os.walk(file_path):
            # file_set = file_set | set(fs)
            file_list.extend(fs)
    except:
        pass
        # return (file_set)
    return (file_list)


def check_patch(file_path=None):
    """
    补丁检查，包括：
    1、待验证补丁文档记录的dll、sql脚本、报表等，与实际版本中的文件能对上（包括修改时间）；
    2、待验证补丁文档，与签入文档内容要对上；
    """
    file_list = get_file_list(file_path)
    file_index = file_list.find('待验证补丁文档.xlsx')
    excel_data = read_excel(file_path + '/' + file_list[file_index])
    # 遍历表格，替换字符串，方便去重
    # 遍历每个工作簿
    for key in excel_data:
        # 遍历每个单元格
        for cell in range(len(excel_data[key])):
            excel_data[cell] = data_format(excel_data[cell])

    # 通过关键字找到表格中记录dll及脚本等文件的列
    # 此处有个局限：严重依赖表格中的文字描述
    for i in excel_data:
        pass
    dll_trait = "对应DLL文件"
    sql_trait = "SQL脚本/报表/其它配置文件(含路径)"
    dll_index = 0
    sql_index = 3


    # 鏁寸悊绋嬪簭銆佽剼鏈�泦
    sql_name_set = set()
    dll_name_set = set()
    for t in dll_set:
        if not t:
            continue
        str_tmp = str(t).replace("\\\\","/")
        str_tmp = str_tmp.replace('\\','/')
        str_tmp = str_tmp.strip()
        if str_tmp[-4:-3] != '.':
            str_tmp += '.dll'
        dll_name_set.add(str_tmp.split('/')[-1])
    for t in sql_set:
        if not t:
            continue
        str_tmp = str(t).replace("\\\\","/")
        str_tmp = str_tmp.replace('\\','/')
        str_tmp = str_tmp.strip()
        if str_tmp[-4:-3] != '.' and str_tmp[-4:] not in ('.sql', '.rps', '.xml'):
            str_tmp += '.sql'
        sql_name_set.add(str_tmp.split('/')[-1])

    # 鎶婃枃妗ｅ悕绉拌浆鎴愬垪琛�紝鏂逛究缁熻�鎵句笉鍒扮殑鏁版嵁
    dll_list = list(dll_name_set)
    sql_list = list(sql_name_set)

    # 閬嶅巻鐗堟湰鐩�綍
    file_list = set()
    for fpathe, dirs, fs in os.walk(file_path):
        file_list = file_list | set(fs)

    for i in file_list:
        file_name = i.lower()
        if file_name[-4:] != '.dll' and file_name[-4:] != '.sql':
            continue

        # 璇诲彇鍒扮殑鏂囦欢涓庢枃妗ｄ腑鐨勬枃浠跺仛閬嶅巻姣斿�
        for dll_tmp in dll_name_set:
            if file_name.lower() == dll_tmp.lower():
                dll_list.pop(dll_list.index(dll_tmp.lower()))
            else:
                pass
        for sql_tmp in sql_name_set:
            if file_name.lower() == sql_tmp.lower():
                sql_list.pop(sql_list.index(sql_tmp.lower()))
            else:
                pass
    if dll_list:
        print('dll miss ---> %s', dll_list)
    if sql_list:
        print('sql miss ===> %s', sql_list)

def check_dll_sql(excel_file, file_path):
    """
    补丁检查，包括：
    1、待验证补丁文档记录的dll、sql脚本、报表等，与实际版本中的文件能对上（包括修改时间）；
    2、待验证补丁文档，与签入文档内容要对上；
    """
    # 在文件中提取 dll、脚本名称，通过集合去重
    dll_set = set()
    sql_set = set()
    workbook = load_workbook(excel_file)
    workbook_sheets_list = workbook.get_sheet_names()
    dll_trait = '对应DLL文件'
    sql_trait = 'SQL脚本/报表/其它配置文件(含路径)'
    dll_index = 0
    sql_index = 3
    for i in workbook_sheets_list:
        sheet = workbook.get_sheet_by_name(i)
        # 遍历工作簿所有单元格
        row_flag = 1
        for j in sheet.rows:
            # 第一行是标题栏
            if row_flag == 1:
                for k in range(len(j)):
                    if j[k].value == dll_trait:
                        dll_index = k
                    elif j[k].value == sql_trait:
                        sql_index = k
            else:
                # 剔除空白行
                if j[dll_index].value:
                    dll_tmp = str(j[dll_index].value).lower()
                    dll_tmp = dll_tmp.split('\n')
                    dll_set |= set(dll_tmp)
                if j[sql_index].value:
                    sql_tmp = str(j[sql_index].value).lower()
                    sql_tmp = sql_tmp.split('\n')
                    sql_set |= set(sql_tmp)
            row_flag += 1

    # 整理程序、脚本集
    sql_name_set = set()
    dll_name_set = set()
    for t in dll_set:
        if not t:
            continue
        str_tmp = str(t).replace("\\\\","/")
        str_tmp = str_tmp.replace('\\','/')
        str_tmp = str_tmp.strip()
        if str_tmp[-4:-3] != '.':
            str_tmp += '.dll'
        dll_name_set.add(str_tmp.split('/')[-1])
    for t in sql_set:
        if not t:
            continue
        str_tmp = str(t).replace("\\\\","/")
        str_tmp = str_tmp.replace('\\','/')
        str_tmp = str_tmp.strip()
        if str_tmp[-4:-3] != '.' and str_tmp[-4:] not in ('.sql', '.rps', '.xml'):
            str_tmp += '.sql'
        sql_name_set.add(str_tmp.split('/')[-1])

    # 把文档名称转成列表，方便统计找不到的数据
    dll_list = list(dll_name_set)
    sql_list = list(sql_name_set)

    # 遍历版本目录
    file_list = set()
    for fpathe, dirs, fs in os.walk(file_path):
        file_list = file_list | set(fs)

    for i in file_list:
        file_name = i.lower()
        if file_name[-4:] != '.dll' and file_name[-4:] != '.sql':
            continue

        # 读取到的文件与文档中的文件做遍历比对
        for dll_tmp in dll_name_set:
            if file_name.lower() == dll_tmp.lower():
                dll_list.pop(dll_list.index(dll_tmp.lower()))
            else:
                pass
        for sql_tmp in sql_name_set:
            if file_name.lower() == sql_tmp.lower():
                sql_list.pop(sql_list.index(sql_tmp.lower()))
            else:
                pass
    if dll_list:
        print('dll miss ---> %s', dll_list)
    if sql_list:
        print('sql miss ===> %s', sql_list)

def read_excel_for_patch(excel_file=None, output='new.xlsx'):
    """
    根据待验证补丁文档，整理出一份新功能列表
    """
    # 新建一个文件，用来存放输出结果
    wb_new = Workbook()
    # 新建一张表
    ws_new = wb_new.active
    # 新增一行表头
    ws_new.append(['版本新功能列表'])
    ws_new.append(["序号","类型","系统模块","恒康需求编号","客户需求编号","涉及的客户","功能名称/修改说明","对现有业务的影响","备注"])

    try:
        # 打开文件
        workbook = load_workbook(excel_file)
        # 获取所有sheet，返回列表，格式：[u'sheet1', u'sheet2']
        workbook_sheets_list = workbook.get_sheet_names()
        # print(workbook_sheets_list)
        # 根据特定字段来查找需要的内容
        bugid_trait = '问题/需求编号'
        modification_trait = '功能/问题修改说明'
        bugid_index = 1
        modification_index = 2
        for i in workbook_sheets_list:
            sheet = workbook.get_sheet_by_name(i)
            # 遍历工作簿所有单元格
            row_flag = 1
            for j in sheet.rows:
                # 第一行是标题栏
                if row_flag == 1:
                    for k in range(len(j)):
                        if j[k].value == bugid_trait:
                            bugid_index = k
                        elif j[k].value == modification_trait:
                            modification_index = k
                else:
                    # 剔除空白行
                    if j[bugid_index].value or j[modification_index].value:
                        bug_id = ''
                        demand_id = ''
                        bug_or_demand = str(j[bugid_index].value)
                        bug_or_demand = bug_or_demand.strip()
                        if bug_or_demand != 'None':
                            if str(bug_or_demand)[0:1].isalpha() and str(bug_or_demand)[0:3].lower() != 'bug':
                                demand_id = bug_or_demand
                            else:
                                bug_id = bug_or_demand
                        ws_new.append(['',bug_id,sheet.title,demand_id,'','',j[modification_index].value,'',''])
                row_flag +=1
    except  Exception as bug:
        ws_new.append(['we have a problme. i have a bug'])
        print(bug)
        # print(j[bugid_index].value)
    finally:
        try:
            wb_new.save(output)
        except  Exception as err:
            print('无法保存文件，文件可能正在被编辑')
    return

def get_demand_from_excel(path=None, output='demand.xlsx'):
    """
    功能：遍历指定文件夹，把全部的xlsx文件查找一次，把有需求编号的都列出来
    返回：成功true，失败False
    """



if __name__ == '__main__':

    excel_path = r'./'
    out_path = './'
    file_list = get_file_list("./")
    print(file_list)
    for i in file_list:
        file_name = i.upper()
        if file_name[-5:] == '.XLSX':
            excel_file = excel_path + file_name
            out_file = out_path + file_name + '产品新功能列表说明.xlsx'
            read_excel_for_patch(excel_file, out_file)
    """
    # os.system("explorer " + out_path)
    # excel_data = read_excel(excel_file)
    # print(excel_data['杩涢攢瀛�][0].index('淇�敼鏃ユ湡'))
    # file = r'D:\jobs\澶栧彂鐗堟湰\閫氱敤\V1.25\V1.25.18.001(2017-6-14)\V1.25.18.001(2017-6-14)琛ヤ竵璇存槑鏂囨。.xlsx'
    # check_path = r'\\Hk-office-fs01\鍝佽川绠″埗閮╛鍐呴儴鏂囦欢$\瀹㈡埛鍗囩骇鐗堟湰\V1.28鐗堟湰\琛ヤ竵\鏈嶈�鐗堟湰\寰呴獙璇佽ˉ涓�
    # file = check_path + '\V1.28鏈嶈�寰呴獙璇佽ˉ涓佹枃妗�xlsx'
    # check_dll_sql(file,check_path)
   
    # excel_file = r"D:\jobs\AutoTest\auto_patch\test_case\V1.27.19.001(2017-12-13)补丁说明文档.xlsx"
    # work_path = r"D:\jobs\AutoTest\auto_patch\test_case"    
    work_path = r"D:\jobs\外发版本\优美\v1.28\优美_V1.28.11.001(2017-12-26)"
    dll_mi_path, dll_mi_excel, sql_mi_path, sql_mi_excel = patch_check(work_path)
    #print("===在excel表格中没有找到的程序文件：")
    #print(dll_mi_path)
    #print("===在本地没有找到的程序文件：")
    #print(dll_mi_excel)
    #print("===在excel表格中没有找到的脚本文件：")
    #print(sql_mi_path)
    #print("===在本地没有找到的脚本文件：")
    #print(sql_mi_excel)
    """